import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Listbox, MULTIPLE
import pandas as pd
import os
import glob
from pathlib import Path
import threading
from datetime import datetime
from openpyxl import Workbook
import hashlib
import sys


class LogFileMerger:
    def __init__(self, root):
        self.root = root
        self.root.title("日志文件合并工具 - 高级版")
        self.root.geometry("950x700")

        # 设置样式
        style = ttk.Style()
        style.theme_use('clam')

        # 初始化变量
        self.file_paths = []
        self.total_files = 0
        self.processed_files = 0
        self.default_save_path = os.path.expanduser("~/Documents")

        # 日志类型选项
        self.log_types = ['elog', 'hwlog', 'trx_status', 'csread', 'vsread', 'tsread', 'all']
        self.selected_log_types = ['all']  # 默认全选

        # 创建GUI组件
        self.create_widgets()

    def create_widgets(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 标题
        title_label = ttk.Label(main_frame, text="日志文件合并工具 - 高级版",
                                font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 10))

        # 文件选择部分
        file_frame = ttk.LabelFrame(main_frame, text="选择文件或文件夹", padding="10")
        file_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))

        # 单文件选择按钮
        ttk.Button(file_frame, text="选择Excel文件",
                   command=self.select_files).grid(row=0, column=0, padx=(0, 10))

        # 文件夹选择按钮
        ttk.Button(file_frame, text="选择文件夹",
                   command=self.select_folder).grid(row=0, column=1, padx=(0, 10))

        # 清空选择按钮
        ttk.Button(file_frame, text="清空选择",
                   command=self.clear_selection).grid(row=0, column=2)

        # 文件列表显示
        list_frame = ttk.LabelFrame(main_frame, text="已选择的文件", padding="10")
        list_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))

        # 添加滚动条
        list_scrollbar = ttk.Scrollbar(list_frame)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.file_listbox = tk.Listbox(list_frame, height=8,
                                       yscrollcommand=list_scrollbar.set,
                                       selectmode=tk.EXTENDED)
        self.file_listbox.pack(fill=tk.BOTH, expand=True)
        list_scrollbar.config(command=self.file_listbox.yview)

        # 配置权重
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)

        # 过滤选项部分
        filter_frame = ttk.LabelFrame(main_frame, text="日志过滤选项", padding="10")
        filter_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))

        # 日志类型选择
        ttk.Label(filter_frame, text="选择日志类型:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))

        # 日志类型列表框
        self.log_type_listbox = Listbox(filter_frame, height=5, selectmode=MULTIPLE, exportselection=False)
        self.log_type_listbox.grid(row=0, column=1, rowspan=2, sticky=(tk.W, tk.E), padx=(0, 10))

        # 添加日志类型选项
        for log_type in self.log_types:
            self.log_type_listbox.insert(tk.END, log_type)

        # 默认选中"all"
        for i, log_type in enumerate(self.log_types):
            if log_type == 'all':
                self.log_type_listbox.selection_set(i)
                break

        # 日志类型选择按钮
        ttk.Button(filter_frame, text="全选",
                   command=self.select_all_log_types).grid(row=0, column=2, padx=(0, 10), sticky=tk.W)
        ttk.Button(filter_frame, text="清空",
                   command=self.clear_log_types).grid(row=0, column=3, sticky=tk.W)

        # 去重选项
        self.remove_duplicates_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(filter_frame, text="去除重复日志条目",
                        variable=self.remove_duplicates_var).grid(row=1, column=2, columnspan=2, sticky=tk.W)

        # 进度显示
        progress_frame = ttk.Frame(main_frame)
        progress_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame,
                                            variable=self.progress_var,
                                            maximum=100)
        self.progress_bar.pack(fill=tk.X, expand=True)

        self.status_label = ttk.Label(progress_frame, text="准备就绪")
        self.status_label.pack()

        # 输出选项
        options_frame = ttk.LabelFrame(main_frame, text="输出选项", padding="10")
        options_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))

        # 第一行：输出文件名和格式
        ttk.Label(options_frame, text="输出文件名:").grid(row=0, column=0, sticky=tk.W)
        self.output_name_var = tk.StringVar(value="merged_logs")
        ttk.Entry(options_frame, textvariable=self.output_name_var, width=25).grid(row=0, column=1, padx=(5, 0))

        ttk.Label(options_frame, text="输出格式:").grid(row=0, column=2, padx=(20, 5))
        self.format_var = tk.StringVar(value="excel")
        ttk.Radiobutton(options_frame, text="Excel", variable=self.format_var,
                        value="excel").grid(row=0, column=3, padx=(0, 10))
        ttk.Radiobutton(options_frame, text="CSV", variable=self.format_var,
                        value="csv").grid(row=0, column=4)

        # 第二行：保存路径设置
        ttk.Label(options_frame, text="保存路径:").grid(row=1, column=0, sticky=tk.W, pady=(10, 0))

        self.save_path_var = tk.StringVar(value=self.default_save_path)
        self.save_path_entry = ttk.Entry(options_frame, textvariable=self.save_path_var, width=35)
        self.save_path_entry.grid(row=1, column=1, columnspan=3, pady=(10, 0), padx=(5, 0), sticky=(tk.W, tk.E))

        ttk.Button(options_frame, text="浏览",
                   command=self.browse_save_path).grid(row=1, column=4, pady=(10, 0), padx=(5, 0))

        # 第三行：保存选项
        self.ask_save_location_var = tk.BooleanVar(value=True)
        self.ask_save_location_check = ttk.Checkbutton(
            options_frame,
            text="合并后询问保存位置",
            variable=self.ask_save_location_var,
            command=self.toggle_save_location_option
        )
        self.ask_save_location_check.grid(row=2, column=0, columnspan=5, pady=(10, 0), sticky=tk.W)

        # 配置列权重
        options_frame.columnconfigure(1, weight=1)

        # 操作按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=6, column=0, columnspan=3, pady=(10, 0))

        ttk.Button(button_frame, text="开始合并",
                   command=self.start_merge).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Button(button_frame, text="退出",
                   command=self.root.quit).pack(side=tk.LEFT)

        # 日志显示
        log_frame = ttk.LabelFrame(main_frame, text="操作日志", padding="10")
        log_frame.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))

        # 配置网格权重
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(7, weight=1)

        log_scrollbar = ttk.Scrollbar(log_frame)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.log_text = tk.Text(log_frame, height=8, wrap=tk.WORD,
                                yscrollcommand=log_scrollbar.set,
                                state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        log_scrollbar.config(command=self.log_text.yview)

        # 设置标签样式
        self.log_text.tag_config("INFO", foreground="blue")
        self.log_text.tag_config("SUCCESS", foreground="green")
        self.log_text.tag_config("WARNING", foreground="orange")
        self.log_text.tag_config("ERROR", foreground="red")

        # 配置filter_frame列权重
        filter_frame.columnconfigure(1, weight=1)

    def select_all_log_types(self):
        """选择所有日志类型"""
        self.log_type_listbox.selection_clear(0, tk.END)
        for i in range(self.log_type_listbox.size()):
            self.log_type_listbox.selection_set(i)
        self.log_message("已选择所有日志类型", "INFO")

    def clear_log_types(self):
        """清空日志类型选择"""
        self.log_type_listbox.selection_clear(0, tk.END)
        self.log_message("已清空日志类型选择", "INFO")

    def toggle_save_location_option(self):
        """切换是否询问保存位置"""
        if self.ask_save_location_var.get():
            self.save_path_entry.config(state='normal')
            self.log_message("保存模式: 合并后将询问保存位置", "INFO")
        else:
            self.save_path_entry.config(state='disabled')
            self.log_message("保存模式: 将自动保存到指定路径", "INFO")

    def browse_save_path(self):
        """浏览保存路径"""
        folder = filedialog.askdirectory(title="选择保存路径", initialdir=self.save_path_var.get())
        if folder:
            self.save_path_var.set(folder)
            self.log_message(f"保存路径已设置为: {folder}", "INFO")

    def log_message(self, message, level="INFO"):
        """在日志区域添加消息"""
        self.log_text.config(state=tk.NORMAL)
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", level)
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update()

    def select_files(self):
        """选择多个Excel文件"""
        files = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if files:
            self.add_files(files)

    def select_folder(self):
        """选择文件夹并自动查找符合条件的文件"""
        folder = filedialog.askdirectory(title="选择文件夹")

        if folder:
            self.find_files_in_folder(folder)

    def find_files_in_folder(self, folder_path):
        """在文件夹中查找所有以_extracted结尾的Excel文件"""
        pattern = os.path.join(folder_path, "**", "*_extracted.xlsx")
        files = glob.glob(pattern, recursive=True)

        pattern2 = os.path.join(folder_path, "**", "*_extracted.xls")
        files.extend(glob.glob(pattern2, recursive=True))

        if files:
            self.add_files(files)
        else:
            messagebox.showwarning("未找到文件",
                                   f"在文件夹 '{folder_path}' 中未找到以 '_extracted' 结尾的Excel文件")

    def add_files(self, files):
        """添加文件到列表"""
        new_files = []
        for file in files:
            if file not in self.file_paths:
                self.file_paths.append(file)
                new_files.append(file)

        if new_files:
            self.update_file_list()
            self.log_message(f"添加了 {len(new_files)} 个文件", "SUCCESS")
        else:
            self.log_message("没有新文件添加", "WARNING")

    def update_file_list(self):
        """更新文件列表框"""
        self.file_listbox.delete(0, tk.END)
        for file in self.file_paths:
            # 显示文件名和部分路径
            display_name = os.path.basename(file)
            parent_dir = os.path.basename(os.path.dirname(file))
            self.file_listbox.insert(tk.END, f"{display_name} (在 {parent_dir} 中)")

    def clear_selection(self):
        """清空选择的文件"""
        self.file_paths.clear()
        self.file_listbox.delete(0, tk.END)
        self.log_message("已清空文件列表", "INFO")

    def start_merge(self):
        """开始合并文件"""
        if not self.file_paths:
            messagebox.showwarning("没有文件", "请先选择要合并的文件或文件夹")
            return

        # 获取选择的日志类型
        selected_indices = self.log_type_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("没有选择日志类型", "请选择至少一种日志类型")
            return

        self.selected_log_types = [self.log_type_listbox.get(i) for i in selected_indices]

        # 检查是否选择了"all"，如果是则包含所有类型
        if 'all' in self.selected_log_types:
            self.selected_log_types = self.log_types[:-1]  # 排除"all"选项

        self.log_message(f"选择的日志类型: {', '.join(self.selected_log_types)}", "INFO")

        # 验证保存路径
        if not self.ask_save_location_var.get():
            save_path = self.save_path_var.get().strip()
            if not save_path or not os.path.exists(save_path):
                messagebox.showerror("保存路径无效", "请指定一个有效的保存路径")
                return

        # 在新线程中执行合并操作，避免界面卡顿
        thread = threading.Thread(target=self.merge_files)
        thread.daemon = True
        thread.start()

    def merge_files(self):
        """合并文件的主要逻辑"""
        try:
            self.total_files = len(self.file_paths)
            self.processed_files = 0

            all_data = []
            excel_max_rows = 1048576  # Excel最大行数
            processed_entries = 0
            skipped_entries = 0
            serial_mismatch_count = 0
            log_type_mismatch_count = 0

            for file_path in self.file_paths:
                try:
                    self.status_label.config(text=f"正在处理: {os.path.basename(file_path)}")

                    # 从文件路径中提取产品串号
                    path_parts = Path(file_path).parts
                    product_serial_from_path = None

                    # 查找产品串号（最后一个文件夹名）
                    if path_parts:
                        product_serial_from_path = path_parts[-2]  # 假设产品串号是文件所在目录

                    # 从文件路径中提取Issue类型
                    issue_type = None
                    for i, part in enumerate(path_parts):
                        if 'issue' in part.lower() or 'problem' in part.lower():
                            issue_type = part
                            break

                    # 如果没找到，使用倒数第三级目录
                    if not issue_type and len(path_parts) >= 3:
                        issue_type = path_parts[-3]

                    # 读取Excel文件
                    df = pd.read_excel(file_path)

                    # 检查必要的列是否存在
                    required_columns = ['Serial', 'ProductName', 'log_type', 'log_line']
                    missing_columns = [col for col in required_columns if col not in df.columns]

                    if missing_columns:
                        self.log_message(f"文件 {os.path.basename(file_path)} 缺少列: {missing_columns}", "WARNING")
                        continue

                    # 创建处理后的数据框
                    filtered_data = []

                    # 处理每一行数据
                    for index, row in df.iterrows():
                        # 检查产品串号是否匹配
                        if product_serial_from_path and str(row['Serial']) != str(product_serial_from_path):
                            serial_mismatch_count += 1
                            continue

                        # 检查日志类型是否在选择的类型中
                        if row['log_type'] not in self.selected_log_types:
                            log_type_mismatch_count += 1
                            continue

                        # 创建新的行数据
                        filtered_row = {
                            'Serial': row['Serial'],
                            'ProductName': row['ProductName'],
                            'log_type': row['log_type'],
                            'log_line': row['log_line'],
                            'issue_type': issue_type if issue_type else 'Unknown',
                            'source_file': os.path.basename(file_path)  # 添加源文件名用于跟踪
                        }

                        filtered_data.append(filtered_row)
                        processed_entries += 1

                    # 如果有筛选后的数据，添加到总数据中
                    if filtered_data:
                        filtered_df = pd.DataFrame(filtered_data)
                        all_data.append(filtered_df)

                    self.processed_files += 1
                    progress = (self.processed_files / self.total_files) * 100
                    self.progress_var.set(progress)

                    self.log_message(f"已处理: {os.path.basename(file_path)} - 保留 {len(filtered_data)} 条记录",
                                     "SUCCESS")

                except Exception as e:
                    self.log_message(f"处理文件 {file_path} 时出错: {str(e)}", "ERROR")

            if not all_data:
                self.log_message("没有有效数据可合并", "ERROR")
                return

            # 合并所有数据
            self.status_label.config(text="正在合并数据...")
            merged_df = pd.concat(all_data, ignore_index=True)

            # 去除重复的日志条目（基于log_line）
            if self.remove_duplicates_var.get():
                initial_count = len(merged_df)
                # 创建一个hash列用于去重
                merged_df['log_line_hash'] = merged_df['log_line'].apply(
                    lambda x: hashlib.md5(str(x).encode()).hexdigest()[:10]
                )
                # 去除重复，保留第一条
                merged_df = merged_df.drop_duplicates(subset=['log_line_hash'], keep='first')
                merged_df = merged_df.drop(columns=['log_line_hash'])
                duplicate_count = initial_count - len(merged_df)
                self.log_message(f"已去除 {duplicate_count} 条重复记录", "SUCCESS")

            # 检查数据量
            total_rows = len(merged_df)

            # 输出处理统计信息
            self.log_message(f"合并完成，统计信息:", "SUCCESS")
            self.log_message(f"  总处理文件数: {self.total_files}", "INFO")
            self.log_message(f"  保留的日志条目数: {total_rows}", "INFO")
            self.log_message(f"  串号不匹配跳过的条目数: {serial_mismatch_count}", "INFO")
            self.log_message(f"  日志类型不匹配跳过的条目数: {log_type_mismatch_count}", "INFO")
            self.log_message(f"  总处理的条目数: {processed_entries + serial_mismatch_count + log_type_mismatch_count}",
                             "INFO")

            # 根据输出格式保存数据
            if self.format_var.get() == "excel":
                self.save_to_excel_with_multiple_sheets(merged_df)
            else:
                self.save_to_csv(merged_df)

            self.status_label.config(text="合并完成")
            self.progress_var.set(100)

        except Exception as e:
            self.log_message(f"合并过程中出错: {str(e)}", "ERROR")
            messagebox.showerror("错误", f"合并过程中出错:\n{str(e)}")

    def save_to_excel_with_multiple_sheets(self, df):
        """将数据保存到Excel，如果数据超过单个工作表最大行数，则创建多个工作表"""
        try:
            output_name = self.output_name_var.get().strip()
            if not output_name:
                output_name = "merged_logs"

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_max_rows = 1048576  # Excel最大行数

            # 检查是否需要询问保存位置
            if self.ask_save_location_var.get():
                file_path = filedialog.asksaveasfilename(
                    title="保存合并文件",
                    defaultextension=".xlsx",
                    filetypes=[("Excel文件", "*.xlsx")],
                    initialfile=f"{output_name}_{timestamp}"
                )

                if not file_path:
                    self.log_message("用户取消了保存操作", "WARNING")
                    return
            else:
                # 自动保存到指定路径
                output_dir = self.save_path_var.get().strip()
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir)

                file_path = os.path.join(output_dir, f"{output_name}_{timestamp}.xlsx")

            self.status_label.config(text="正在保存到Excel文件...")

            # 计算需要的工作表数量
            total_rows = len(df)
            num_sheets = (total_rows // excel_max_rows) + 1

            # 使用openpyxl创建Excel写入器
            wb = Workbook()

            # 删除默认创建的工作表
            default_sheet = wb.active
            wb.remove(default_sheet)

            # 将数据分割到多个工作表中
            for sheet_num in range(num_sheets):
                start_idx = sheet_num * excel_max_rows
                end_idx = min((sheet_num + 1) * excel_max_rows, total_rows)

                # 获取当前工作表的数据
                chunk_df = df.iloc[start_idx:end_idx]

                # 创建工作表
                sheet_name = f"Sheet{sheet_num + 1}"
                if sheet_num == 0:
                    ws = wb.create_sheet(title=sheet_name)
                else:
                    ws = wb.create_sheet(title=sheet_name)

                # 写入表头
                for col_num, column_name in enumerate(chunk_df.columns, 1):
                    ws.cell(row=1, column=col_num, value=column_name)

                # 写入数据行
                for row_num, row in enumerate(chunk_df.itertuples(index=False), 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num, value=value)

                self.log_message(f"工作表 '{sheet_name}' 已创建，包含 {len(chunk_df)} 行数据", "SUCCESS")

            # 保存Excel文件
            wb.save(file_path)

            self.log_message(f"Excel文件已保存: {file_path}", "SUCCESS")
            self.log_message(f"共创建了 {num_sheets} 个工作表", "INFO")
            messagebox.showinfo("完成",
                                f"Excel文件已保存到:\n{file_path}\n"
                                f"总行数: {total_rows}\n"
                                f"工作表数量: {num_sheets}")

        except Exception as e:
            raise Exception(f"保存Excel文件时出错: {str(e)}")

    def save_to_csv(self, df):
        """将数据保存到CSV，如果数据超过Excel最大行数，则创建多个CSV文件"""
        try:
            output_name = self.output_name_var.get().strip()
            if not output_name:
                output_name = "merged_logs"

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_max_rows = 1048576  # Excel最大行数

            total_rows = len(df)

            # 检查是否需要分割文件（CSV格式不支持多个工作表）
            if total_rows > excel_max_rows:
                # 需要分割为多个CSV文件
                self.split_and_save_csv(df, excel_max_rows)
            else:
                # 保存为单个CSV文件
                if self.ask_save_location_var.get():
                    file_path = filedialog.asksaveasfilename(
                        title="保存合并文件",
                        defaultextension=".csv",
                        filetypes=[("CSV文件", "*.csv")],
                        initialfile=f"{output_name}_{timestamp}"
                    )

                    if not file_path:
                        self.log_message("用户取消了保存操作", "WARNING")
                        return
                else:
                    # 自动保存到指定路径
                    output_dir = self.save_path_var.get().strip()
                    if not os.path.exists(output_dir):
                        os.makedirs(output_dir)

                    file_path = os.path.join(output_dir, f"{output_name}_{timestamp}.csv")

                self.status_label.config(text="正在保存CSV文件...")

                df.to_csv(file_path, index=False, encoding='utf-8-sig')

                self.log_message(f"CSV文件已保存: {file_path}", "SUCCESS")
                messagebox.showinfo("完成", f"CSV文件已保存到:\n{file_path}\n总行数: {total_rows}")

        except Exception as e:
            raise Exception(f"保存CSV文件时出错: {str(e)}")

    def split_and_save_csv(self, df, max_rows):
        """分割数据并保存到多个CSV文件（仅用于CSV格式）"""
        try:
            num_files = (len(df) // max_rows) + 1
            output_name = self.output_name_var.get().strip()
            if not output_name:
                output_name = "merged_logs"

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # 检查是否需要询问保存位置
            if self.ask_save_location_var.get():
                output_dir = filedialog.askdirectory(title="选择保存位置")
                if not output_dir:
                    self.log_message("用户取消了保存操作", "WARNING")
                    return
            else:
                output_dir = self.save_path_var.get().strip()
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir)

            self.status_label.config(text=f"正在分割数据到 {num_files} 个CSV文件...")

            for i in range(num_files):
                start_idx = i * max_rows
                end_idx = min((i + 1) * max_rows, len(df))

                chunk_df = df.iloc[start_idx:end_idx]

                output_path = os.path.join(output_dir, f"{output_name}_{timestamp}_part{i + 1}.csv")
                chunk_df.to_csv(output_path, index=False, encoding='utf-8-sig')

                self.log_message(f"已保存: {os.path.basename(output_path)} ({len(chunk_df)} 行)", "SUCCESS")

            messagebox.showinfo("完成",
                                f"数据已分割保存到 {num_files} 个CSV文件中\n保存位置: {output_dir}")

        except Exception as e:
            raise Exception(f"保存CSV文件时出错: {str(e)}")


def main():
    root = tk.Tk()
    app = LogFileMerger(root)

    # 使窗口可调整大小
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)

    # 设置窗口图标
    try:
        root.iconbitmap(default='icon.ico')
    except:
        pass

    root.mainloop()


if __name__ == "__main__":
    main()
